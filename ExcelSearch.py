import os
import multiprocessing
import pandas
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

from openpyxl import load_workbook
from xlrd import open_workbook
from colorama import Fore, Style

def red_text(text):
    return Fore.RED+Style.BRIGHT+text+Style.RESET_ALL

def blue_text(text):
    return Fore.BLUE+Style.BRIGHT+text+Style.RESET_ALL

def yellow_text(text):
    return Fore.YELLOW+Style.BRIGHT+text+Style.RESET_ALL

def convert(input_file, output_file):
    pandas.read_excel(input_file).to_excel(output_file, index=False)

PATH = './temp' # main.py 也存在

def search_xls(file_path, search_term):
    wb = open_workbook(file_path)

    for sheet in wb.sheets():
        for row in range(sheet.nrows):
            for col in range(sheet.ncols):
                cell_value = str(sheet.cell_value(row, col)).lower().strip()
                if search_term in cell_value:
                    return True
    return False

def single_search(file_path, search_term, original_path=None):
    if file_path.endswith('.xls'):
        try:
            if not os.path.exists(PATH):
                os.mkdir(PATH)

            if search_xls(file_path, search_term):
                file_name = os.path.basename(file_path) # xxx.xls
                output_file = f'{PATH}/{file_name}x'
                convert(file_path, output_file)
                single_search(output_file, search_term, file_path)

        except Exception as e:
            print(yellow_text(f"[Search1 Error]{file_path}: {e}\n"))
        return

    try:
        if original_path is None:
            original_path = file_path

        workbook = load_workbook(file_path, read_only=True)  # 使用只读模式提高性能
        results = []

        for sheet in workbook.sheetnames:
            data = {'rmb': None, 'usd': None, 'part': None, 'description': None}
            data2 = {'part': "件号", 'description': "品名", 'rmb': "RMB", 'usd': "USD"}

            worksheet = workbook[sheet]
            for row in worksheet.iter_rows():
                for cell in row:
                    cell_value = str(cell.value).lower().strip()
                    # 检测需要打印的数据在哪列
                    for i in data:
                        if data[i] is None and i in cell_value:
                            data[i] = cell.coordinate[0]

                    if search_term in cell_value:
                        results.append(f'在 "{original_path}" "{sheet}" "{cell.coordinate}" 找到目标')
                        output_parts = []
                        for i in data:
                            if data[i] is not None:
                                index = data[i]+cell.coordinate[1:]
                                value = str(worksheet[index].value)
                                colored = red_text(value) if i in ('usd', 'rmb') else blue_text(value)
                                output_parts.append(f"{data2[i]}: {colored}")

                        results.append(f"  {yellow_text('|')}  ".join(output_parts) + "\n")

        for result in results:
            print(result)

    except Exception as e:
        print(yellow_text(f"[Search2 Error]{original_path}: {e}\n"))

    finally:
        if 'workbook' in locals():
            workbook.close()

def batch_search(folder_path, data_to_find, max_workers=8):
    search_term = data_to_find.lower().strip()

    file_paths = [
        os.path.join(root, file).replace('\\', '/')
        for root, _, files in os.walk(folder_path)
        for file in files
        if file.endswith(('.xlsx', '.xls')) and not file.startswith(('~$', '$'))
    ]

    with multiprocessing.Pool(processes=max_workers) as pool:
        pool.starmap(single_search, [(fp, search_term) for fp in file_paths])

if __name__ == '__main__':
    batch_search(r'D:\stress test', 'ACTUATOR  8431499900')